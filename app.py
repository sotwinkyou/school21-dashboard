import streamlit as st
import pandas as pd
import requests
import urllib.parse
from docx import Document
from icalendar import Calendar
from datetime import datetime
import io
import time
import zipfile

# --- ПРЕДОТВРАЩЕНИЕ ОШИБОК СТИЛЕЙ (MONKEYPATCH OPENPYXL) ---
try:
    import openpyxl.reader.excel
    orig_read_style = openpyxl.reader.excel.ExcelReader.read_style
    def safe_read_style(self):
        try:
            orig_read_style(self)
        except Exception:
            # Безопасные дефолты при критической ошибке XML-стилей Яндекса
            self.shared_styles = []
            self.stylesheet = None
    openpyxl.reader.excel.ExcelReader.read_style = safe_read_style
except Exception:
    pass

import openpyxl

# Настройки страницы Streamlit
st.set_page_config(
    page_title="Дашборд Школы 21", 
    layout="wide", 
    page_icon="💚"
)

# Очистка и запуск автоматического обновления (каждые 120 секунд)
if "last_update" not in st.session_state:
    st.session_state.last_update = time.time()

current_time = time.time()
if current_time - st.session_state.last_update > 120:
    st.session_state.last_update = current_time
    st.cache_data.clear()

# Фирменный стиль Сбера и Школы 21 (бело-зеленая гамма)
st.markdown("""
    <style>
    /* Основной фон приложения */
    .stApp {
        background-color: #f4f6f8;
    }
    
    /* Стилизация заголовков */
    h1 {
        color: #08a652 !important;
        font-family: 'SB Sans Text', 'Helvetica Neue', sans-serif;
        font-weight: 800;
        border-bottom: 3px solid #08a652;
        padding-bottom: 10px;
    }
    
    h2, h3 {
        color: #111827;
        font-family: 'SB Sans Text', sans-serif;
        font-weight: 600;
    }
    
    /* Стилизация карточек KPI */
    div[data-testid="stMetricValue"] {
        font-size: 2rem !important;
        font-weight: bold;
        color: #08a652 !important;
    }
    
    div[data-testid="stMetricLabel"] {
        font-size: 0.95rem !important;
        color: #4b5563 !important;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    
    /* Боковая панель */
    section[data-testid="stSidebar"] {
        background-color: #ffffff;
        border-right: 1px solid #e5e7eb;
    }
    
    /* Кнопки и интерактивные элементы */
    .stButton>button {
        background-color: #08a652 !important;
        color: white !important;
        border-radius: 8px !important;
        border: none !important;
        font-weight: 600 !important;
        transition: all 0.3s ease;
        width: 100%;
    }
    
    .stButton>button:hover {
        background-color: #06803f !important;
        box-shadow: 0 4px 12px rgba(8, 166, 82, 0.2);
    }
    
    /* Красивые контейнеры для KPI */
    .kpi-card {
        background-color: #ffffff;
        padding: 20px;
        border-radius: 12px;
        border-left: 5px solid #08a652;
        box-shadow: 0 2px 8px rgba(0,0,0,0.04);
        margin-bottom: 15px;
    }
    </style>
""", unsafe_allow_html=True)

st.title("💚 Автоматизированный дашборд | Школа 21")

# --- ФУНКЦИЯ ЗАГРУЗКИ С ЯНДЕКС.ДИСКА ---
@st.cache_data(ttl=120)
def download_from_yandex(public_url):
    """
    Скачивание файла по публичной ссылке Яндекс.Диска
    """
    base_url = 'https://cloud-api.yandex.net/v1/disk/public/resources/download'
    final_url = base_url + '?public_key=' + urllib.parse.quote(public_url)
    try:
        response = requests.get(final_url)
        if response.status_code == 200:
            download_url = response.json().get('href')
            file_response = requests.get(download_url)
            if file_response.status_code == 200:
                return file_response.content, None
            else:
                return None, f"Не удалось скачать файл. Код: {file_response.status_code}"
        else:
            return None, "Нет доступа. Проверьте, что ссылка публичная."
    except Exception as e:
        return None, f"Ошибка подключения к Яндексу: {str(e)}"

# --- ХИРУРГИЧЕСКОЕ УДАЛЕНИЕ СТИЛЕЙ ИЗ XLSX-АРХИВА ---
def strip_styles_from_xlsx(file_bytes):
    """
    Удаляет файл стилей xl/styles.xml из zip-архива xlsx.
    Это на 100% решает проблему с поломанными XML-стилями Яндекса.
    """
    try:
        in_mem_zip = io.BytesIO(file_bytes)
        out_mem_zip = io.BytesIO()
        with zipfile.ZipFile(in_mem_zip, 'r') as yin:
            with zipfile.ZipFile(out_mem_zip, 'w', zipfile.ZIP_DEFLATED) as yout:
                for item in yin.infolist():
                    if item.filename == 'xl/styles.xml':
                        continue
                    yout.writestr(item, yin.read(item.filename))
        return out_mem_zip.getvalue()
    except Exception:
        return file_bytes

# --- ФУНКЦИИ УМНОГО ПАРСИНГА ФАЙЛОВ ---
def parse_excel(file_bytes):
    """
    Парсинг Excel с предварительной очисткой от проблемных стилей.
    Поддерживает чтение нескольких листов.
    """
    try:
        sanitized_bytes = strip_styles_from_xlsx(file_bytes)
        wb = openpyxl.load_workbook(io.BytesIO(sanitized_bytes), read_only=True, data_only=True)
        
        # Словарь для хранения данных со всех листов
        sheets_data = {}
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            data = []
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    data.append(list(row))
            
            if not data:
                continue
                
            headers = [str(h).strip() if h is not None else f"Col_{i}" for i, h in enumerate(data[0])]
            df = pd.DataFrame(data[1:], columns=headers)
            
            # Умный маппинг колонок
            column_mapping = {}
            for col in df.columns:
                col_lower = col.lower()
                if "назван" in col_lower or "событи" in col_lower or "мероприяти" in col_lower:
                    column_mapping[col] = "Название"
                elif "дат" in col_lower:
                    column_mapping[col] = "Дата"
                elif "врем" in col_lower or "час" in col_lower:
                    column_mapping[col] = "Время"
                elif "мест" in col_lower or "аудитор" in col_lower or "кластер" in col_lower:
                    column_mapping[col] = "Место"
                elif "участн" in col_lower or "кол-во" in col_lower or "люд" in col_lower:
                    column_mapping[col] = "Участники"
                    
            df = df.rename(columns=column_mapping)
            
            # Если нашли базовые колонки событий, сохраняем как главную таблицу
            required_cols = ["Название", "Дата", "Время", "Место", "Участники"]
            has_events = any(col in df.columns for col in ["Название", "Дата"])
            
            if has_events:
                for rc in required_cols:
                    if rc not in df.columns:
                        df[rc] = "Не указано"
                sheets_data[sheet_name] = df[required_cols]
            else:
                # Иначе сохраняем лист как вспомогательный (например, список студентов)
                sheets_data[sheet_name] = df
                
        # Если нашли листы, собираем события
        event_sheets = [df for name, df in sheets_data.items() if "Название" in df.columns]
        if event_sheets:
            final_df = pd.concat(event_sheets, ignore_index=True)
            return final_df, sheets_data, None
        elif sheets_data:
            # Возвращаем первый попавшийся лист, если события не распознались по колонкам
            first_name = list(sheets_data.keys())[0]
            return sheets_data[first_name], sheets_data, None
            
        return None, {}, "Файл не содержит данных."
    except Exception as e:
        return None, {}, f"Ошибка Excel: {str(e)}"

def parse_docx(file_bytes):
    try:
        doc = Document(io.BytesIO(file_bytes))
        data = []
        for table in doc.tables:
            for row in table.rows[1:]:
                text = [cell.text.strip() for cell in row.cells]
                if len(text) >= 5:
                    data.append(text[:5])
        df = pd.DataFrame(data, columns=["Название", "Дата", "Время", "Место", "Участники"])
        return df, None
    except Exception as e:
        return None, f"Ошибка Word: {str(e)}"

def parse_ical(file_bytes):
    try:
        gcal = Calendar.from_ical(file_bytes)
        data = []
        for component in gcal.walk():
            if component.name == "VEVENT":
                summary = str(component.get('summary'))
                start = component.get('dtstart').dt
                location = str(component.get('location', 'Кампус'))
                
                date_str = start.strftime("%Y-%m-%d") if isinstance(start, datetime) else str(start)
                time_str = start.strftime("%H:%M") if isinstance(start, datetime) else "00:00"
                
                data.append([summary, date_str, time_str, location, "Не указано"])
        df = pd.DataFrame(data, columns=["Название", "Дата", "Время", "Место", "Участники"])
        return df, None
    except Exception as e:
        return None, f"Ошибка Календаря: {str(e)}"

# --- БОКОВАЯ ПАНЕЛЬ И НАСТРОЙКИ СВЯЗИ ---
st.sidebar.image("https://upload.wikimedia.org/wikipedia/commons/e/e0/Sberbank_Logo_2020.svg", width=120)
st.sidebar.header("📁 Параметры импорта")

query_params = st.query_params
default_table_url = query_params.get("table", "")
default_cal_url = query_params.get("cal", "")

source_type = st.sidebar.radio("Источник данных:", ["Яндекс.Диск ссылки", "Локальные файлы"])

all_events = pd.DataFrame(columns=["Название", "Дата", "Время", "Место", "Участники"])
additional_sheets = {}
errors = []

if source_type == "Локальные файлы":
    uploaded_files = st.sidebar.file_uploader(
        "Загрузите файлы (Excel, Word, iCal)", 
        type=["xlsx", "docx", "ics"], 
        accept_multiple_files=True
    )
    if uploaded_files:
        for file in uploaded_files:
            file_bytes = file.read()
            if file.name.endswith('.xlsx'):
                df, sheets, err = parse_excel(file_bytes)
                if sheets:
                    additional_sheets.update(sheets)
            elif file.name.endswith('.docx'):
                df, err = parse_docx(file_bytes)
            elif file.name.endswith('.ics'):
                df, err = parse_ical(file_bytes)
            
            if err:
                errors.append(f"Ошибка в файле {file.name}: {err}")
            elif df is not None:
                all_events = pd.concat([all_events, df], ignore_index=True)

else:
    st.sidebar.info("💡 Данные будут автоматически синхронизироваться раз в 2 минуты.")
    yandex_url_1 = st.sidebar.text_input("Ссылка на Таблицу (Excel) или Документ (Word):", value=default_table_url, placeholder="https://disk.yandex.ru/i/...")
    yandex_url_2 = st.sidebar.text_input("Ссылка на Календарь (.ics) [опционально]:", value=default_cal_url, placeholder="https://disk.yandex.ru/d/...")
    
    if yandex_url_1 or yandex_url_2:
        st.query_params["table"] = yandex_url_1
        st.query_params["cal"] = yandex_url_2

    urls_to_process = []
    if yandex_url_1: urls_to_process.append((yandex_url_1, "table"))
    if yandex_url_2: urls_to_process.append((yandex_url_2, "cal"))
    
    if urls_to_process:
        for url, u_type in urls_to_process:
            file_bytes, err = download_from_yandex(url)
            if err:
                errors.append(f"Ошибка загрузки по ссылке ({u_type}): {err}")
            elif file_bytes:
                if ".xlsx" in url or "excel" in url.lower() or u_type == "table":
                    df, sheets, err = parse_excel(file_bytes)
                    if sheets:
                        additional_sheets.update(sheets)
                elif ".docx" in url or "word" in url.lower():
                    df, err = parse_docx(file_bytes)
                elif ".ics" in url or "calendar" in url.lower() or u_type == "cal":
                    df, err = parse_ical(file_bytes)
                else:
                    df, sheets, err = parse_excel(file_bytes)
                    if sheets:
                        additional_sheets.update(sheets)
                
                if err:
                    errors.append(f"Ошибка распознавания данных: {err}")
                elif df is not None:
                    all_events = pd.concat([all_events, df], ignore_index=True)

    if st.sidebar.button("🔄 Синхронизировать сейчас"):
        st.cache_data.clear()
        st.rerun()

# --- ДИНАМИЧЕСКИЙ РАСЧЕТ KPI ИЗ ДАННЫХ ---
students_count = "432"
active_projects = "18"
deadline_info = "2 дня"
loading_staff = "Ветка С & DevOps"

# Пытаемся автоматически найти KPI во вспомогательных листах Excel
for sheet_name, df_sheet in additional_sheets.items():
    sheet_str = df_sheet.to_string().lower()
    if "студент" in sheet_str or "пир" in sheet_str:
        # Если нашли колонку с ID пиров или студентов, считаем их количество
        students_count = str(len(df_sheet))
    if "проект" in sheet_str:
        active_projects = str(len(df_sheet))
    if "дедлайн" in sheet_str or "срок" in sheet_str:
        deadline_info = "Ближайший"

# Если в основной таблице есть участники, попробуем просуммировать их
if not all_events.empty and "Участники" in all_events.columns:
    try:
        # Пытаемся извлечь числа из колонки участников
        only_nums = all_events["Участники"].astype(str).str.extract(r'(\d+)').dropna()
        if not only_nums.empty:
            total_ppl = only_nums[0].astype(int).sum()
            loading_staff = f"{total_ppl} чел. на ивентах"
    except Exception:
        pass

# --- ГЛАВНЫЙ ЭКРАН ---

# 1. БЛОК KPI
st.subheader("🚀 Пульс Кампуса")
kpi_col1, kpi_col2, kpi_col3, kpi_col4 = st.columns(4)

with kpi_col1:
    st.markdown('<div class="kpi-card">', unsafe_allow_html=True)
    st.metric(label="👥 Студенты (Сheck-in)", value=students_count, delta="Синхронизировано" if len(additional_sheets) > 0 else "+12 сегодня")
    st.markdown('</div>', unsafe_allow_html=True)
with kpi_col2:
    st.markdown('<div class="kpi-card">', unsafe_allow_html=True)
    st.metric(label="💻 Активные проекты", value=active_projects, delta="Из загруженных файлов" if len(additional_sheets) > 0 else "Ветка С & DevOps")
    st.markdown('</div>', unsafe_allow_html=True)
with kpi_col3:
    st.markdown('<div class="kpi-card">', unsafe_allow_html=True)
    st.metric(label="📅 Всего мероприятий", value=len(all_events) if not all_events.empty else "12")
    st.markdown('</div>', unsafe_allow_html=True)
with kpi_col4:
    st.markdown('<div class="kpi-card">', unsafe_allow_html=True)
    st.metric(label="⏳ Сроки и дедлайны", value=deadline_info, delta="Финал Бассейна", delta_color="inverse")
    st.markdown('</div>', unsafe_allow_html=True)

st.markdown("---")

# 2. РАСПИСАНИЕ И УМНАЯ ФИЛЬТРАЦИЯ ПО МЕСЯЦАМ
st.subheader("🗓️ Расписание и график событий")

if not all_events.empty:
    # Оповещение об успешной загрузке событий
    st.success(f"🎉 Успешно синхронизировано событий: {len(all_events)}")
    
    # Нормализация дат для умного парсинга периодов
    all_events['parsed_date'] = pd.to_datetime(all_events['Дата'], errors='coerce')
    
    # Создание списка месяцев на русском языке
    RU_MONTHS = {
        1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель", 5: "Май", 6: "Июнь",
        7: "Июль", 8: "Август", 9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
    }
    
    # Генерация опций для фильтра по месяцам
    filter_options = ["Предстоящие события", "Все время (включая прошлые)"]
    
    valid_dates = all_events['parsed_date'].dropna()
    if not valid_dates.empty:
        # Извлекаем уникальные Год-Месяц
        unique_periods = sorted(list(set((d.year, d.month) for d in valid_dates)), key=lambda x: (x[0], x[1]), reverse=True)
        for year, month in unique_periods:
            filter_options.append(f"{RU_MONTHS[month]} {year}")
            
    # Виджеты фильтрации на главном экране
    col_f1, col_f2 = st.columns([2, 1])
    with col_f1:
        selected_period = st.selectbox("📅 Выберите период отображения (включая прошедшие месяцы):", options=filter_options)
    with col_f2:
        # Фильтр по локациям
        all_events['Место'] = all_events['Место'].fillna('Не указано')
        places = all_events['Место'].unique()
        selected_place = st.multiselect("📍 Локации (Кластеры):", options=places, default=places)
    
    # Применение временного фильтра
    current_date = datetime.now() # Июнь 2026
    
    if selected_period == "Предстоящие события":
        # Показываем только будущие или сегодняшние события
        filtered_df = all_events[(all_events['parsed_date'] >= current_date) | (all_events['parsed_date'].isna())]
    elif selected_period == "Все время (включая прошлые)":
        filtered_df = all_events
    else:
        # Фильтр по конкретному выбранному Месяцу Года
        month_name, year_str = selected_period.split()
        target_year = int(year_str)
        target_month = [k for k, v in RU_MONTHS.items() if v == month_name][0]
        
        filtered_df = all_events[
            (all_events['parsed_date'].dt.year == target_year) & 
            (all_events['parsed_date'].dt.month == target_month)
        ]
        
    # Применение фильтра по локациям
    filtered_df = filtered_df[filtered_df['Место'].isin(selected_place)]
    
    # Показ таблицы с данными
    if not filtered_df.empty:
        # Убираем временную техническую колонку перед показом
        show_df = filtered_df.drop(columns=['parsed_date']) if 'parsed_date' in filtered_df.columns else filtered_df
        st.dataframe(show_df, use_container_width=True)
    else:
        st.warning("В выбранном периоде или локациях события отсутствуют.")
        
    # ЕСЛИ загружены дополнительные листы (например, студенты, проекты) - выводим их во вкладках!
    other_sheets = [name for name in additional_sheets.keys() if name not in event_sheets] if 'event_sheets' in locals() else list(additional_sheets.keys())
    if other_sheets:
        st.write("---")
        st.subheader("📊 Дополнительные таблицы из файлов")
        tabs = st.tabs(other_sheets)
        for i, tab_name in enumerate(other_sheets):
            with tabs[i]:
                st.dataframe(additional_sheets[tab_name], use_container_width=True)

else:
    st.info("Используйте боковую панель для загрузки локальных файлов или подключения ссылок Яндекс.Диска.")
    st.caption("Пример отображения расписания (демо-данные):")
    demo_data = pd.DataFrame({
        "Название": ["Хакатон Сбера", "Peer-to-Peer Защиты", "Встреча с HR", "Лекция по AI"],
        "Дата": ["2026-07-01", "2026-07-02", "2026-07-03", "2026-07-04"],
        "Время": ["10:00", "14:00", "16:30", "12:00"],
        "Место": ["Конференц-зал", "Кластер А", "Переговорка 2", "Кластер Б"],
        "Участники": ["120 человек", "Все пиры", "Команда Core", "Бассейн"]
    })
    st.dataframe(demo_data, use_container_width=True)

st.markdown("---")

# 3. ОТЧЕТЫ И БЛОКЕРЫ
st.subheader("📋 Оперативные отчеты")
col_left, col_right = st.columns(2)

with col_left:
    st.write("**📊 Прогресс подготовки к ивентам:**")
    st.progress(0.8, text="Организация Летнего Хакатона (80%)")
    st.progress(0.4, text="Закупка мерча для бассейна (40%)")

with col_right:
    st.write("**⚠️ Контроль блокеров и импорта:**")
    if errors:
        for err in errors:
            st.error(err)
    else:
        st.success("Все импортированные файлы синхронизированы без ошибок.")
